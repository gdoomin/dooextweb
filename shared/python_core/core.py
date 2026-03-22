import json
import re
import unicodedata
from math import atan2, cos, radians, sin, sqrt
from pathlib import Path
from typing import Any


POLYGON_ONLY_MESSAGE = "폴리곤(도형) 파일입니다. 시작점/끝점 추출 대상이 아닙니다."
_GENERIC_LINE_NAMES = {"line", "linestring", "flight line"}
EARTH_RADIUS_KM = 6371.0088
AIRCRAFT_SPEED_KNOTS = 130
KNOT_TO_KMH = 1.852
TURN_MINUTES_PER_LINE = 3


def wlen(value: str) -> int:
    return sum(2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1 for ch in value)


def rjust_w(value: str, width: int) -> str:
    return " " * max(0, width - wlen(value)) + value


def center_w(value: str, width: int) -> str:
    pad = max(0, width - wlen(value))
    left = pad // 2
    return " " * left + value + " " * (pad - left)


def dd_to_dms(dd: float, is_lat: bool = True) -> str:
    direction = ("N" if dd >= 0 else "S") if is_lat else ("E" if dd >= 0 else "W")
    total_centiseconds = int(round(abs(dd) * 3600 * 100))
    degree, remaining_centiseconds = divmod(total_centiseconds, 3600 * 100)
    minute, second_centiseconds = divmod(remaining_centiseconds, 60 * 100)
    second = second_centiseconds / 100
    return f"{degree}\u00B0{minute:02d}'{second:05.2f}\"{direction}"


def _read_kml(filepath: str) -> str | None:
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return Path(filepath).read_text(encoding=encoding)
        except (UnicodeDecodeError, LookupError):
            continue
    return None


def _parse_coords(raw: str) -> list[tuple[float, float]]:
    points: list[tuple[float, float]] = []
    for token in re.split(r"\s+", raw.strip()):
        parts = token.split(",")
        if len(parts) < 2:
            continue
        try:
            points.append((float(parts[0]), float(parts[1])))
        except ValueError:
            continue
    return points


def _normalize_polygon_points(points: list[tuple[float, float]]) -> list[tuple[float, float]]:
    if len(points) >= 2 and points[0] == points[-1]:
        return points[:-1]
    return points


def _get_root(content: str):
    import xml.etree.ElementTree as ET

    clean = re.sub(r'\sxmlns[^=]*="[^"]*"', "", content)
    clean = re.sub(r"<([/]?)\w+:", r"<\1", clean)
    return ET.fromstring(clean)


def _iter_placemarks_with_context(node, folder_names: tuple[str, ...] = ()):
    current_folders = folder_names
    if node.tag == "Folder":
        folder_name = (node.findtext("name") or "").strip()
        if folder_name:
            current_folders = folder_names + (folder_name,)

    if node.tag == "Placemark":
        yield node, current_folders
        return

    for child in list(node):
        yield from _iter_placemarks_with_context(child, current_folders)


def _extract_context_line_name(folder_names: tuple[str, ...]) -> str:
    for folder_name in reversed(folder_names):
        name = (folder_name or "").strip()
        if not name:
            continue

        match = re.search(r"(?:flight\s*line|line)\s*\[([^\]]+)\]", name, re.IGNORECASE)
        if match:
            return match.group(1).strip()

        match = re.search(r"\[([^\]]+)\]", name)
        if match:
            return match.group(1).strip()

        if re.fullmatch(r"[A-Za-z]?\d+[A-Za-z]?", name):
            return name

    return ""


def _resolve_linestring_name(name: str, folder_names: tuple[str, ...]) -> str:
    clean_name = (name or "").strip()
    normalized = re.sub(r"\s+", " ", clean_name).lower()
    if clean_name and normalized not in _GENERIC_LINE_NAMES:
        return clean_name

    context_name = _extract_context_line_name(folder_names)
    return context_name or clean_name


def _parse_linestring(filepath: str):
    content = _read_kml(filepath)
    if not content:
        return None, "파일 인코딩을 읽을 수 없습니다."

    results: list[dict[str, Any]] = []
    try:
        root = _get_root(content)
        for placemark, folder_names in _iter_placemarks_with_context(root):
            raw_name = (placemark.findtext("name") or "").strip()
            name = _resolve_linestring_name(raw_name, folder_names)
            line_string = placemark.find(".//LineString")
            if line_string is None:
                continue
            coords_node = line_string.find(".//coordinates")
            if coords_node is None or not coords_node.text:
                continue
            points = _parse_coords(coords_node.text)
            if len(points) >= 2:
                results.append(
                    {
                        "num": name,
                        "s_lat": points[0][1],
                        "s_lon": points[0][0],
                        "e_lat": points[-1][1],
                        "e_lon": points[-1][0],
                    }
                )
    except Exception:
        pass

    if not results:
        for block in re.findall(r"<[^>]*Placemark[^>]*>(.*?)</[^>]*Placemark>", content, re.DOTALL | re.IGNORECASE):
            name_match = re.search(r"<name[^>]*>(.*?)</name>", block, re.DOTALL | re.IGNORECASE)
            raw_name = re.sub(r"<[^>]+>", "", name_match.group(1)).strip() if name_match else ""
            coords_match = re.search(
                r"<[^>]*LineString[^>]*>.*?<[^>]*coordinates[^>]*>(.*?)</[^>]*coordinates>",
                block,
                re.DOTALL | re.IGNORECASE,
            )
            if not coords_match:
                continue
            points = _parse_coords(coords_match.group(1))
            if len(points) >= 2:
                results.append(
                    {
                        "num": raw_name,
                        "s_lat": points[0][1],
                        "s_lon": points[0][0],
                        "e_lat": points[-1][1],
                        "e_lon": points[-1][0],
                    }
                )

    if not results:
        return None, None
    return results, None


def _parse_polygon(filepath: str):
    content = _read_kml(filepath)
    if not content:
        return None, "파일 인코딩을 읽을 수 없습니다."

    polygons: list[dict[str, Any]] = []
    try:
        root = _get_root(content)
        for placemark in root.iter("Placemark"):
            base_name = (placemark.findtext("name") or "").strip()
            polygon_nodes = list(placemark.findall(".//Polygon"))
            for index, polygon in enumerate(polygon_nodes, 1):
                coord_nodes = polygon.findall(".//outerBoundaryIs//coordinates") or polygon.findall(".//coordinates")
                if not coord_nodes:
                    continue
                coords_node = coord_nodes[0]
                if coords_node is None or not coords_node.text:
                    continue
                points = _normalize_polygon_points(_parse_coords(coords_node.text))
                if len(points) >= 3:
                    name = base_name
                    if base_name and len(polygon_nodes) > 1:
                        name = f"{base_name} #{index}"
                    polygons.append({"num": name, "points": points})
    except Exception:
        pass

    if not polygons:
        for block in re.findall(r"<[^>]*Placemark[^>]*>(.*?)</[^>]*Placemark>", content, re.DOTALL | re.IGNORECASE):
            name_match = re.search(r"<name[^>]*>(.*?)</name>", block, re.DOTALL | re.IGNORECASE)
            base_name = re.sub(r"<[^>]+>", "", name_match.group(1)).strip() if name_match else ""
            coord_blocks = re.findall(
                r"<[^>]*Polygon[^>]*>.*?<[^>]*coordinates[^>]*>(.*?)</[^>]*coordinates>",
                block,
                re.DOTALL | re.IGNORECASE,
            )
            for index, coords in enumerate(coord_blocks, 1):
                points = _normalize_polygon_points(_parse_coords(coords))
                if len(points) >= 3:
                    name = base_name
                    if base_name and len(coord_blocks) > 1:
                        name = f"{base_name} #{index}"
                    polygons.append({"num": name, "points": points})

    if not polygons:
        return None, "LineString와 Polygon을 찾을 수 없습니다."
    return polygons, None


def parse_kml(filepath: str):
    results, error = _parse_linestring(filepath)
    if results is not None:
        return results, None, "linestring"
    if error is not None:
        return None, error, None

    results, error = _parse_polygon(filepath)
    if results is not None:
        return results, None, "polygon"
    return None, error or "데이터를 찾을 수 없습니다.", None


def format_text(results: list[dict[str, Any]], project_name: str = "", mode: str = "linestring") -> str:
    col_line = 4
    col_kind = 4
    col_lat = 22
    col_lon = 23
    sep = "  "

    if mode == "polygon":
        summary = f"폴리곤 수: {len(results)}개"
        if project_name:
            return f"프로젝트: {project_name}\n{'=' * 70}\n{POLYGON_ONLY_MESSAGE}\n{summary}"
        return f"{POLYGON_ONLY_MESSAGE}\n{summary}"

    lines: list[str] = []
    if project_name:
        lines.append(f"프로젝트: {project_name}")
        lines.append("=" * 70)

    lines.append(
        rjust_w("Line", col_line)
        + sep
        + rjust_w("구분", col_kind)
        + sep
        + center_w("위도", col_lat)
        + sep
        + center_w("경도", col_lon)
    )
    lines.append("-" * 70)

    for row in results:
        lines.append(
            rjust_w(str(row.get("num", "")), col_line)
            + sep
            + rjust_w("시작", col_kind)
            + sep
            + center_w(dd_to_dms(row["s_lat"], True), col_lat)
            + sep
            + center_w(dd_to_dms(row["s_lon"], False), col_lon)
        )
        lines.append(
            " " * col_line
            + sep
            + rjust_w("끝", col_kind)
            + sep
            + center_w(dd_to_dms(row["e_lat"], True), col_lat)
            + sep
            + center_w(dd_to_dms(row["e_lon"], False), col_lon)
        )
        lines.append("")

    return "\n".join(lines)


def save_excel(
    results: list[dict[str, Any]],
    filepath: str,
    project_name: str = "",
    mode: str = "linestring",
    polygons: list[dict[str, Any]] | None = None,
) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = openpyxl.Workbook()
    default_sheet = workbook.active

    title_font = Font(name="맑은 고딕", bold=True, size=11)
    header_font = Font(name="맑은 고딕", bold=True, size=10)
    body_font = Font(name="맑은 고딕", size=10)
    gray_fill = PatternFill("solid", fgColor="F2F2F2")
    title_fill = PatternFill("solid", fgColor="D9D9D9")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    thin_side = Side(style="thin", color="999999")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def to_float(value: Any) -> float | None:
        try:
            number = float(value)
        except (TypeError, ValueError):
            return None
        if number != number or number in (float("inf"), float("-inf")):
            return None
        return number

    def normalize_polygon_rows() -> list[dict[str, Any]]:
        normalized: list[dict[str, Any]] = []
        if isinstance(polygons, list) and polygons:
            for index, row in enumerate(polygons, 1):
                if not isinstance(row, dict):
                    continue
                raw_points = row.get("points")
                if not isinstance(raw_points, list):
                    continue
                points_latlon: list[tuple[float, float]] = []
                for point in raw_points:
                    if not isinstance(point, (list, tuple)) or len(point) < 2:
                        continue
                    lat = to_float(point[0])
                    lon = to_float(point[1])
                    if lat is None or lon is None:
                        continue
                    points_latlon.append((lat, lon))
                if len(points_latlon) < 3:
                    continue
                label = str(row.get("label") or row.get("num") or f"Polygon {index}").strip() or f"Polygon {index}"
                normalized.append(
                    {
                        "label": label,
                        "count": len(points_latlon),
                        "first_lat": points_latlon[0][0],
                        "first_lon": points_latlon[0][1],
                    }
                )
            return normalized

        if mode == "polygon":
            for index, row in enumerate(results, 1):
                if not isinstance(row, dict):
                    continue
                raw_points = row.get("points")
                if not isinstance(raw_points, list):
                    continue
                points_latlon: list[tuple[float, float]] = []
                for point in raw_points:
                    if not isinstance(point, (list, tuple)) or len(point) < 2:
                        continue
                    lon = to_float(point[0])
                    lat = to_float(point[1])
                    if lat is None or lon is None:
                        continue
                    points_latlon.append((lat, lon))
                if len(points_latlon) < 3:
                    continue
                label = str(row.get("num") or f"Polygon {index}").strip() or f"Polygon {index}"
                normalized.append(
                    {
                        "label": label,
                        "count": len(points_latlon),
                        "first_lat": points_latlon[0][0],
                        "first_lon": points_latlon[0][1],
                    }
                )
        return normalized

    def write_line_sheet() -> None:
        sheet = default_sheet if default_sheet.title == "Sheet" else workbook.create_sheet("코스좌표")
        sheet.title = "코스좌표"

        sheet.merge_cells("A1:F1")
        title = project_name.strip() if project_name.strip() else "코스 좌표"
        sheet.cell(row=1, column=1, value=title)

        headers = ["번호", "라인", "시작 위도", "시작 경도", "끝 위도", "끝 경도"]
        for column, header in enumerate(headers, 1):
            sheet.cell(row=2, column=column, value=header)

        row_index = 3
        for idx, row in enumerate(results, 1):
            s_lat = to_float(row.get("s_lat")) if isinstance(row, dict) else None
            s_lon = to_float(row.get("s_lon")) if isinstance(row, dict) else None
            e_lat = to_float(row.get("e_lat")) if isinstance(row, dict) else None
            e_lon = to_float(row.get("e_lon")) if isinstance(row, dict) else None
            if s_lat is None or s_lon is None or e_lat is None or e_lon is None:
                continue
            label = str(row.get("num", "")).strip() if isinstance(row, dict) else ""
            sheet.cell(row=row_index, column=1, value=idx)
            sheet.cell(row=row_index, column=2, value=label or "-")
            sheet.cell(row=row_index, column=3, value=dd_to_dms(s_lat, True))
            sheet.cell(row=row_index, column=4, value=dd_to_dms(s_lon, False))
            sheet.cell(row=row_index, column=5, value=dd_to_dms(e_lat, True))
            sheet.cell(row=row_index, column=6, value=dd_to_dms(e_lon, False))
            row_index += 1

        for row in sheet.iter_rows(min_row=1, max_row=max(2, row_index - 1), min_col=1, max_col=6):
            for cell in row:
                if cell.row == 1:
                    cell.font = title_font
                    cell.fill = title_fill
                elif cell.row == 2:
                    cell.font = header_font
                    cell.fill = gray_fill
                else:
                    cell.font = body_font
                    cell.fill = white_fill
                cell.alignment = center
                cell.border = border

        widths = {"A": 8, "B": 12, "C": 20, "D": 20, "E": 20, "F": 20}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

    def write_polygon_sheet(rows: list[dict[str, Any]]) -> None:
        sheet = default_sheet if default_sheet.title == "Sheet" else workbook.create_sheet("폴리곤요약")
        sheet.title = "폴리곤요약"

        sheet.merge_cells("A1:F1")
        title = project_name.strip() if project_name.strip() else "폴리곤 요약"
        sheet.cell(row=1, column=1, value=title)

        headers = ["번호", "이름", "포인트 수", "대표 위도", "대표 경도", "안내"]
        for column, header in enumerate(headers, 1):
            sheet.cell(row=2, column=column, value=header)

        row_index = 3
        for idx, row in enumerate(rows, 1):
            sheet.cell(row=row_index, column=1, value=idx)
            sheet.cell(row=row_index, column=2, value=row["label"])
            sheet.cell(row=row_index, column=3, value=row["count"])
            sheet.cell(row=row_index, column=4, value=dd_to_dms(float(row["first_lat"]), True))
            sheet.cell(row=row_index, column=5, value=dd_to_dms(float(row["first_lon"]), False))
            sheet.cell(row=row_index, column=6, value=POLYGON_ONLY_MESSAGE)
            row_index += 1

        if row_index == 3:
            sheet.cell(row=3, column=1, value="데이터 없음")
            sheet.merge_cells("A3:F3")
            row_index = 4

        for row in sheet.iter_rows(min_row=1, max_row=max(2, row_index - 1), min_col=1, max_col=6):
            for cell in row:
                if cell.row == 1:
                    cell.font = title_font
                    cell.fill = title_fill
                    cell.alignment = left
                elif cell.row == 2:
                    cell.font = header_font
                    cell.fill = gray_fill
                    cell.alignment = center
                else:
                    cell.font = body_font
                    cell.fill = white_fill
                    cell.alignment = left if cell.column in (2, 6) else center
                cell.border = border

        widths = {"A": 8, "B": 24, "C": 12, "D": 20, "E": 20, "F": 52}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

    polygon_rows = normalize_polygon_rows()
    has_line_rows = mode != "polygon" and any(
        isinstance(row, dict) and all(key in row for key in ("s_lat", "s_lon", "e_lat", "e_lon"))
        for row in results
    )

    if has_line_rows:
        write_line_sheet()
    if polygon_rows or mode == "polygon":
        write_polygon_sheet(polygon_rows)

    if default_sheet.title == "Sheet" and len(workbook.sheetnames) > 1:
        workbook.remove(default_sheet)
    elif default_sheet.title == "Sheet" and len(workbook.sheetnames) == 1:
        default_sheet.title = "요약"
        default_sheet.cell(row=1, column=1, value=project_name.strip() or "요약")
        default_sheet.cell(row=2, column=1, value="내보낼 데이터가 없습니다.")

    workbook.save(filepath)


def _build_force_labels(results: list[dict[str, Any]]) -> dict[int, str]:
    ordered = sorted(
        range(len(results)),
        key=lambda idx: (
            -((results[idx]["s_lat"] + results[idx]["e_lat"]) / 2.0),
            ((results[idx]["s_lon"] + results[idx]["e_lon"]) / 2.0),
        ),
    )
    return {idx: str(rank) for rank, idx in enumerate(ordered, 1)}


def _haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    lat1_rad = radians(lat1)
    lat2_rad = radians(lat2)
    delta_lat = radians(lat2 - lat1)
    delta_lon = radians(lon2 - lon1)
    a = sin(delta_lat / 2) ** 2 + cos(lat1_rad) * cos(lat2_rad) * sin(delta_lon / 2) ** 2
    return 2 * EARTH_RADIUS_KM * atan2(sqrt(a), sqrt(1 - a))


def _build_linestring_meta_text(results: list[dict[str, Any]]) -> str:
    total_length_km = 0.0
    for row in results:
        total_length_km += _haversine_km(row["s_lat"], row["s_lon"], row["e_lat"], row["e_lon"])

    flight_hours = total_length_km / (AIRCRAFT_SPEED_KNOTS * KNOT_TO_KMH) if total_length_km > 0 else 0.0
    turn_hours = (len(results) * TURN_MINUTES_PER_LINE) / 60
    total_capture_hours = flight_hours + turn_hours
    return (
        f"{len(results)}\uac1c \ub77c\uc778"
        f" \u00b7 \ucd1d\uae38\uc774 {total_length_km:.1f}km"
        f" \u00b7 \ucd1d\ucd2c\uc601\uc2dc\uac04 : \ub300\ub7b5 {total_capture_hours:.1f}\uc2dc\uac04"
    )


def build_web_map_payload(
    results: list[dict[str, Any]],
    project_name: str,
    mode: str = "linestring",
    layer_catalog: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    layer_catalog = layer_catalog or []

    if mode == "polygon":
        payload_polygons = []
        for index, row in enumerate(results, 1):
            points = [[round(lat, 6), round(lon, 6)] for lon, lat in row.get("points", ())]
            if len(points) < 3:
                continue
            payload_polygons.append(
                {
                    "num": row.get("num", ""),
                    "label": row.get("num", "").strip() or f"Polygon {index}",
                    "points": points,
                }
            )

        return {
            "project_name": project_name,
            "mode": mode,
            "results": [],
            "polygons": payload_polygons,
            "has_kml_num": False,
            "default_force_num": False,
            "default_show_num": False,
            "has_layers": bool(layer_catalog),
            "layer_catalog": layer_catalog,
            "default_gray_map": False,
            "meta_text": f"{len(payload_polygons)}개 폴리곤",
        }

    has_kml_num = bool(results) and all(bool(str(row.get("num", "")).strip()) for row in results)
    force_labels = _build_force_labels(results)
    payload_results = []

    for index, row in enumerate(results):
        payload_results.append(
            {
                "num": row.get("num", ""),
                "force_label": force_labels.get(index, str(index + 1)),
                "force_order": int(force_labels.get(index, str(index + 1))),
                "s_lat": round(row["s_lat"], 6),
                "s_lon": round(row["s_lon"], 6),
                "e_lat": round(row["e_lat"], 6),
                "e_lon": round(row["e_lon"], 6),
                "s_text": f"{dd_to_dms(row['s_lat'], True)} {dd_to_dms(row['s_lon'], False)}",
                "e_text": f"{dd_to_dms(row['e_lat'], True)} {dd_to_dms(row['e_lon'], False)}",
            }
        )

    return {
        "project_name": project_name,
        "mode": mode,
        "results": payload_results,
        "polygons": [],
        "has_kml_num": has_kml_num,
        "default_force_num": not has_kml_num,
        "default_show_num": has_kml_num,
        "has_layers": bool(layer_catalog),
        "layer_catalog": layer_catalog,
        "default_gray_map": False,
        "meta_text": _build_linestring_meta_text(results),
    }


def build_web_map_html(payload: dict[str, Any], template_path: str) -> str:
    template = Path(template_path).read_text(encoding="utf-8")
    payload_json = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    return template.replace("__PAYLOAD_JSON__", payload_json)
